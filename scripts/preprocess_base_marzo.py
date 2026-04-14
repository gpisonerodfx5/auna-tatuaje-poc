"""
Preprocesador BASE_MARZO.xlsx -> CSV para el Parser Lambda.

Aplica las reglas del documento BASE_MARZO_context.md:
  1. Lee hoja '6 a 11' del xlsx
  2. Filtra consentimiento IN ('SI','SI') (acepta con o sin tilde, minus/mayus)
  3. Usa celular_afil_1; si vacio cae a celular_contratante_1
  4. Normaliza telefono a formato +51XXXXXXXXX
  5. Mapea distrito_afil -> centerId (sede_referencia) via tabla hardcoded
  6. Renombra columnas al formato que espera el Parser Lambda:
     - numero_documento_afil, telefono, apellidos_nombres_afil,
       programa_final, sede_referencia, cantidad_cuotas_pagadas,
       grupo_cuota_pagada, cod_campana
  7. Escribe el CSV resultante (local o subido a S3)

Uso:
  python scripts/preprocess_base_marzo.py <input.xlsx> <output.csv>
  python scripts/preprocess_base_marzo.py BASE_MARZO.xlsx dist/afiliados.csv
  python scripts/preprocess_base_marzo.py BASE_MARZO.xlsx --s3 s3://auna-tatuaje-poc-input-769488154338/afiliados.csv

Requiere: openpyxl, boto3 (solo si usas --s3)
  pip install openpyxl
"""

import argparse
import csv
import logging
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
log = logging.getLogger(__name__)

# Tabla de mapeo distrito -> centerId, de BASE_MARZO_context.md
# Fuente: doc + Centers by city (Lima y Callao)
DISTRITO_A_CENTER_ID = {
    # Lima y Callao
    "LIMA": 4,                       # Delgado
    "SANTIAGO DE SURCO": 8,          # OC Benavides
    "SURCO": 8,
    "SAN MARTIN DE PORRES": 10,      # Auna Guardia Civil
    "SAN JUAN DE LURIGANCHO": 18,    # C.B. Independencia
    "COMAS": 18,
    "MIRAFLORES": 11,                # OC San Isidro
    "LOS OLIVOS": 10,
    "SAN BORJA": 11,
    "SAN ISIDRO": 11,
    "SURQUILLO": 11,
    "SAN MIGUEL": 9,                 # OC Encalada
    "LA MOLINA": 8,
    "MAGDALENA DEL MAR": 9,
    "MAGDALENA": 9,
    "ATE": 18,
    "CALLAO": 15,                    # Bellavista
    "VILLA MARIA DEL TRIUNFO": 8,
    "PUEBLO LIBRE": 9,
    "CHORRILLOS": 8,
    "LA VICTORIA": 4,
    "RIMAC": 4,
    "BELLAVISTA": 15,
    "INDEPENDENCIA": 18,
    "BREÑA": 4,
    "BRENA": 4,
    "JESUS MARIA": 9,
    "LINCE": 11,
    "BARRANCO": 11,
    "SAN LUIS": 11,
    "LURIN": 8,
    "PUENTE PIEDRA": 18,
    "CARABAYLLO": 18,
    "VENTANILLA": 15,
    # Otras ciudades (una sede principal por ciudad)
    "AREQUIPA": 1,                   # Vallesur
    "TRUJILLO": 2,                   # Camino Real
    "PIURA": 13,                     # Miraflores (Piura)
    "CHICLAYO": 17,                  # Auna Clinica Chiclayo
}

DEFAULT_CENTER_ID = 4  # Delgado - usado como fallback para distritos no mapeados

# Columnas del xlsx (BASE_MARZO_context.md)
COL_DNI = "numero_documento_afil"
COL_NOMBRE = "apellidos_nombres_afil"
COL_TEL_AFIL = "celular_afil_1"
COL_TEL_CONTRATANTE = "celular_contratante_1"
COL_DISTRITO = "distrito_afil"
COL_PROGRAMA = "programa_final"
COL_CUOTAS = "cantidad_cuotas_pagadas"
COL_GRUPO_CUOTA = "grupo_cuota_pagada"
COL_CONSENTIMIENTO = "consentimiento"
COL_COD_CAMPANA = "cod_campana"


def normalize_consentimiento(val) -> bool:
    if val is None:
        return False
    v = str(val).strip().upper().replace("Í", "I")
    return v == "SI"


def normalize_telefono(val) -> str:
    """Normaliza a formato +51XXXXXXXXX. Devuelve '' si invalido."""
    if val is None:
        return ""
    s = re.sub(r"[^\d+]", "", str(val).strip())
    if not s:
        return ""
    if s.startswith("+"):
        # Ya tiene prefijo internacional
        return s
    # Asume Peru y agrega +51 si faltaba
    if s.startswith("51") and len(s) >= 11:
        return f"+{s}"
    if len(s) == 9:  # celular peruano de 9 digitos
        return f"+51{s}"
    # Invalido
    return ""


def normalize_distrito(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    # Normalizar acentos basicos
    s = s.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    return s


def distrito_a_center_id(distrito: str) -> int:
    d = normalize_distrito(distrito)
    return DISTRITO_A_CENTER_ID.get(d, DEFAULT_CENTER_ID)


def preprocess(input_path: str, sheet_name: str = "6 a 11") -> list[dict]:
    log.info(f"Leyendo {input_path} (hoja '{sheet_name}')")
    wb = load_workbook(input_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        log.error(f"Hoja '{sheet_name}' no encontrada. Hojas disponibles: {wb.sheetnames}")
        return []
    ws = wb[sheet_name]

    # Primera fila = headers
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    log.info(f"Headers: {len(headers)} columnas")

    # Verificar que columnas criticas existen
    required = [COL_DNI, COL_NOMBRE, COL_TEL_AFIL, COL_DISTRITO, COL_PROGRAMA,
                COL_CUOTAS, COL_GRUPO_CUOTA, COL_CONSENTIMIENTO]
    missing = [c for c in required if c not in headers]
    if missing:
        log.error(f"Columnas criticas faltantes: {missing}")
        return []

    idx = {h: i for i, h in enumerate(headers)}

    stats = {"total": 0, "sin_consentimiento": 0, "sin_dni": 0, "sin_telefono": 0, "validos": 0}
    afiliados = []

    for row in rows_iter:
        stats["total"] += 1

        def get(col):
            i = idx.get(col, -1)
            return row[i] if i >= 0 and i < len(row) else None

        # Filtro 1: consentimiento
        if not normalize_consentimiento(get(COL_CONSENTIMIENTO)):
            stats["sin_consentimiento"] += 1
            continue

        # Filtro 2: DNI valido
        dni = str(get(COL_DNI) or "").strip()
        if not dni or len(dni) not in (8, 9, 12):
            stats["sin_dni"] += 1
            continue

        # Telefono con fallback al contratante
        telefono = normalize_telefono(get(COL_TEL_AFIL))
        if not telefono:
            telefono = normalize_telefono(get(COL_TEL_CONTRATANTE))
        if not telefono:
            stats["sin_telefono"] += 1
            continue

        # Mapear distrito a centerId
        distrito = str(get(COL_DISTRITO) or "").strip()
        center_id = distrito_a_center_id(distrito)

        afiliados.append({
            "numero_documento_afil": dni,
            "apellidos_nombres_afil": str(get(COL_NOMBRE) or "").strip(),
            "telefono": telefono,
            "programa_final": str(get(COL_PROGRAMA) or "").strip(),
            "sede_referencia": str(center_id),
            "cantidad_cuotas_pagadas": str(get(COL_CUOTAS) or ""),
            "grupo_cuota_pagada": str(get(COL_GRUPO_CUOTA) or ""),
            "cod_campana": str(get(COL_COD_CAMPANA) or "").strip(),
            # Debug fields (el Parser los ignora)
            "_distrito_original": distrito,
        })
        stats["validos"] += 1

    log.info(f"Stats: {stats}")
    return afiliados


def write_csv(afiliados: list[dict], output_path: str):
    if not afiliados:
        log.warning("No hay afiliados validos para escribir")
        return

    # Columnas que el Parser Lambda espera
    fieldnames = [
        "numero_documento_afil",
        "apellidos_nombres_afil",
        "telefono",
        "programa_final",
        "sede_referencia",
        "cantidad_cuotas_pagadas",
        "grupo_cuota_pagada",
        "cod_campana",
    ]

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(afiliados)
    log.info(f"CSV escrito: {output_path} ({len(afiliados)} filas)")


def upload_s3(local_path: str, s3_uri: str):
    import boto3
    m = re.match(r"s3://([^/]+)/(.+)", s3_uri)
    if not m:
        log.error(f"URI S3 invalida: {s3_uri}")
        return
    bucket, key = m.group(1), m.group(2)
    # Usa el perfil auna-sandbox si esta disponible
    try:
        session = boto3.Session(profile_name="auna-sandbox")
    except Exception:
        session = boto3.Session()
    s3 = session.client("s3", region_name="us-east-1")
    s3.upload_file(local_path, bucket, key)
    log.info(f"Subido a s3://{bucket}/{key}")


def main():
    parser = argparse.ArgumentParser(description="Preprocesa BASE_MARZO.xlsx a CSV para el Parser Lambda")
    parser.add_argument("input", help="Archivo BASE_MARZO.xlsx")
    parser.add_argument("output", nargs="?", default="dist/afiliados.csv", help="CSV de salida (local)")
    parser.add_argument("--sheet", default="6 a 11", help="Hoja del xlsx (default: '6 a 11')")
    parser.add_argument("--s3", help="Sube el CSV a S3. Ej: s3://auna-tatuaje-poc-input-769488154338/afiliados.csv")
    parser.add_argument("--limit", type=int, default=0, help="Limita N afiliados para testing (default: sin limite)")

    args = parser.parse_args()

    if not os.path.exists(args.input):
        log.error(f"Archivo no encontrado: {args.input}")
        sys.exit(1)

    afiliados = preprocess(args.input, sheet_name=args.sheet)

    if args.limit > 0:
        afiliados = afiliados[:args.limit]
        log.info(f"Limitado a {len(afiliados)} afiliados para testing")

    write_csv(afiliados, args.output)

    if args.s3:
        upload_s3(args.output, args.s3)


if __name__ == "__main__":
    main()
